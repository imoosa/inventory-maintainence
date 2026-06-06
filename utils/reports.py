"""
ReportGenerator — produces downloadable Excel and PDF reports.

Dependencies:
  pip install openpyxl reportlab
"""


class ReportGenerator:

    # ------------------------------------------------------------------
    # Inventory Excel Report
    # ------------------------------------------------------------------
    def generate_inventory_excel(self, items, filepath):
        """
        Generates a formatted .xlsx inventory report.
        Columns: Asset Code, Name, Category, Brand, Model, Serial No,
                 Location, Qty, Purchase Cost, Current Value,
                 Condition, Status, Purchase Date, Warranty Expiry,
                 Insurance Expiry, AMC End Date, Supplier
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Inventory Report"

            # -- Header row
            headers = [
                'Asset Code', 'Name', 'Category', 'Sub-Category',
                'Brand', 'Model', 'Serial No.', 'Location',
                'Quantity', 'Purchase Cost (₹)', 'Current Value (₹)',
                'Depreciation %', 'Condition', 'Status',
                'Purchase Date', 'Warranty Expiry', 'Insurance Expiry',
                'AMC Provider', 'AMC End Date', 'Supplier'
            ]

            header_fill = PatternFill(start_color='2D6A4F', end_color='2D6A4F', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            center = Alignment(horizontal='center', vertical='center')
            thin = Side(style='thin', color='CCCCCC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border

            ws.row_dimensions[1].height = 22

            # -- Data rows
            alt_fill = PatternFill(start_color='F0F4F0', end_color='F0F4F0', fill_type='solid')
            for row_idx, item in enumerate(items, start=2):
                fill = alt_fill if row_idx % 2 == 0 else PatternFill()

                def fmt_date(d):
                    return d.strftime('%d-%m-%Y') if d else ''

                supplier_name = item.supplier.name if item.supplier else ''
                cat_name = item.category.name if item.category else ''

                row_data = [
                    item.asset_code, item.name, cat_name, item.sub_category or '',
                    item.brand or '', item.model or '', item.serial_number or '',
                    item.location or '', item.quantity,
                    item.purchase_cost or 0, item.current_value or 0,
                    item.depreciation_rate or 0, item.condition or '',
                    item.status or 'active',
                    fmt_date(item.purchase_date), fmt_date(item.warranty_expiry),
                    fmt_date(item.insurance_expiry),
                    item.amc_provider or '', fmt_date(item.amc_end_date),
                    supplier_name
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')

            # -- Column widths
            col_widths = [14, 28, 16, 14, 14, 14, 16, 18, 8, 16, 14, 12, 12,
                          10, 14, 14, 14, 18, 14, 20]
            for i, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # -- Summary sheet
            ws2 = wb.create_sheet(title="Summary")
            ws2['A1'] = 'Summary'
            ws2['A1'].font = Font(bold=True, size=14)
            ws2['A3'] = 'Total Assets'
            ws2['B3'] = len(items)
            ws2['A4'] = 'Total Purchase Cost (₹)'
            ws2['B4'] = sum(i.purchase_cost or 0 for i in items)
            ws2['A5'] = 'Total Current Value (₹)'
            ws2['B5'] = sum(i.current_value or 0 for i in items)
            ws2['A6'] = 'Active Assets'
            ws2['B6'] = sum(1 for i in items if i.status == 'active')
            ws2['A7'] = 'Retired Assets'
            ws2['B7'] = sum(1 for i in items if i.status == 'retired')
            ws2['A8'] = 'Under Maintenance'
            ws2['B8'] = sum(1 for i in items if i.status == 'under_maintenance')

            for row in ws2.iter_rows(min_row=3, max_row=8, min_col=1, max_col=2):
                for cell in row:
                    cell.border = border

            wb.save(filepath)

        except ImportError:
            raise RuntimeError("openpyxl is required: pip install openpyxl")

    # ------------------------------------------------------------------
    # Maintenance PDF Report
    # ------------------------------------------------------------------
    def generate_maintenance_pdf(self, schedules, filepath):
        """
        Generates a formatted PDF maintenance report.
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                            Paragraph, Spacer)

            doc = SimpleDocTemplate(
                filepath,
                pagesize=landscape(A4),
                rightMargin=1.5 * cm,
                leftMargin=1.5 * cm,
                topMargin=2 * cm,
                bottomMargin=1.5 * cm
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'Title', parent=styles['Title'],
                fontSize=16, spaceAfter=12, textColor=colors.HexColor('#1B4332')
            )
            sub_style = ParagraphStyle(
                'Sub', parent=styles['Normal'],
                fontSize=9, textColor=colors.grey, spaceAfter=16
            )

            elements = []
            from datetime import datetime as dt
            elements.append(Paragraph("Maintenance Schedule Report", title_style))
            elements.append(Paragraph(
                f"Generated: {dt.now().strftime('%d %B %Y %H:%M')}  |  "
                f"Total records: {len(schedules)}",
                sub_style
            ))

            headers = ['Asset Name', 'Asset Code', 'Due Date', 'Last Performed',
                       'Frequency', 'Status', 'Priority', 'Total Cost (₹)', 'Notes']

            data = [headers]
            for s in schedules:
                item_name = s.item.name if s.item else 'N/A'
                asset_code = s.item.asset_code if s.item else 'N/A'
                due_date = s.due_date.strftime('%d-%m-%Y') if s.due_date else ''
                last_done = s.last_performed.strftime('%d-%m-%Y') if s.last_performed else 'Never'
                freq = f"{s.frequency_days}d" if s.frequency_days else 'One-time'
                cost = f"₹{s.total_cost:,.0f}" if s.total_cost else '₹0'
                data.append([
                    item_name, asset_code, due_date, last_done,
                    freq, s.status.title(), s.priority.title(), cost,
                    (s.notes or '')[:40]
                ])

            col_widths = [5.5*cm, 3.5*cm, 2.8*cm, 3*cm, 2.2*cm,
                          2.2*cm, 2.2*cm, 2.8*cm, 5*cm]

            status_colors = {
                'pending': colors.HexColor('#FFF3CD'),
                'overdue': colors.HexColor('#F8D7DA'),
                'completed': colors.HexColor('#D4EDDA'),
                'escalated': colors.HexColor('#F8D7DA'),
            }

            table = Table(data, colWidths=col_widths, repeatRows=1)

            style = TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4332')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [colors.white, colors.HexColor('#F5F5F5')]),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ])

            # Color rows by status
            for row_idx, s in enumerate(schedules, start=1):
                bg = status_colors.get(s.status)
                if bg:
                    style.add('BACKGROUND', (5, row_idx), (5, row_idx), bg)

            table.setStyle(style)
            elements.append(table)
            doc.build(elements)

        except ImportError:
            raise RuntimeError("reportlab is required: pip install reportlab")
